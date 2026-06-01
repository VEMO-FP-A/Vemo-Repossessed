#!/usr/bin/env python3
"""
convert.py — Genera data.json para el dashboard de recuperaciones VEMO
Uso: python3 convert.py "VEMO_Recuperaciones_Template.xlsx"
     python3 convert.py   (busca VEMO_Recuperaciones_Template.xlsx en la carpeta)
"""
import sys, json, glob, os
from datetime import datetime
from collections import defaultdict
import openpyxl

# ── Archivo ───────────────────────────────────────────────────────────────────
FNAME = sys.argv[1] if len(sys.argv) > 1 else "VEMO_Recuperaciones_Template.xlsx"
if not os.path.exists(FNAME):
    matches = glob.glob("VEMO_Recuperaciones*.xlsx")
    FNAME = matches[0] if matches else None
if not FNAME:
    print("ERROR: No se encontró VEMO_Recuperaciones_Template.xlsx"); sys.exit(1)
print(f"Leyendo: {FNAME}")

wb = openpyxl.load_workbook(FNAME, read_only=True, data_only=True)

# ── Helpers ───────────────────────────────────────────────────────────────────
def parse_cos(s):
    if not s or not isinstance(s, str): return None
    p = s.strip().split('-')
    if len(p) != 2: return None
    try: return (int(p[0]), int(p[1]))
    except: return None

def sort_key(s):
    p = parse_cos(s)
    return (p[1], p[0]) if p else (9999, 99)

ABBR = {1:'Jan',2:'Feb',3:'Mar',4:'Apr',5:'May',6:'Jun',
        7:'Jul',8:'Aug',9:'Sep',10:'Oct',11:'Nov',12:'Dec'}

# ── 1. EXPORT sheet ──────────────────────────────────────────────────────────
# Columns (0-indexed): A=0 Cliente, B=1 Sexo, C=2 Contrato, D=3 Estado,
# E=4 Cosecha, F=5 MesEnvio, G=6 MesesTransc, H=7 Estatus,
# I=8 MesRecuperacion, J=9 Plate, K=10 Modelo, L=11 Plataforma, M=12 Meses_Efectivos
ws_exp = wb['Export']
rows_exp = list(ws_exp.iter_rows(values_only=True))

cohort_all  = defaultdict(lambda: {'t':0,'ca':0,'p':0})
plat_totals = defaultdict(int)
vintage_raw = defaultdict(list)
monthly_rec = defaultdict(int)

for r in rows_exp[2:]:  # skip banner row + header row
    cos  = r[4]   # Cosecha
    est  = r[7]   # Estatus
    mrec = r[8]   # Mes recuperación
    plat = r[11]  # Plataforma
    mef  = r[12]  # Meses_Efectivos (calculated col M)

    if not cos or not est or not isinstance(cos, str): continue
    p = parse_cos(cos)
    if not p: continue

    cohort_all[cos]['t'] += 1
    if est == 'Efectivo':
        cohort_all[cos]['ca'] += 1
        if mrec and isinstance(mrec, str):
            monthly_rec[mrec] += 1
        if mef is not None:
            try: vintage_raw[cos].append(int(mef))
            except: pass
    elif est == 'Pendiente':
        cohort_all[cos]['p'] += 1

    if plat: plat_totals[str(plat).strip().upper()] += 1

# ── 2. Build VD (vintage curves) ─────────────────────────────────────────────
VD = {}
for cos in set(cohort_all) | set(vintage_raw):
    n = cohort_all[cos]['t']
    if n == 0: continue
    ml = vintage_raw.get(cos, [])
    VD[cos] = {'n': n, 'c': [round(sum(1 for x in ml if x<=m)/n*100,1) for m in range(37)]}

# ── 3. ORIGINACIONES sheet (simple: col A=Cosecha, col B=N_Orig) ─────────────
ws_ori = wb['Originaciones']
N_ORIG = {}
for r in list(ws_ori.iter_rows(values_only=True))[2:]:  # skip banner + header
    cos = r[0]
    n   = r[1]
    if not cos or not isinstance(cos, str): continue
    if not n or not isinstance(n, (int, float)) or int(n) == 0: continue
    if parse_cos(cos): N_ORIG[cos.strip()] = int(n)

# ── 4. ORIG_REC chart (Jan 2025+) ────────────────────────────────────────────
months_2025 = sorted([k for k in N_ORIG if parse_cos(k) and parse_cos(k)[1]>=2025], key=sort_key)
orig_rec = {
    'labels': [f"{ABBR[parse_cos(m)[0]]} {str(parse_cos(m)[1])[2:]}" for m in months_2025],
    'new_v':  [N_ORIG[m] for m in months_2025],
    'rec':    [monthly_rec.get(m, 0) for m in months_2025]
}

# ── 5. COSECHA_TABLE (2024+) ─────────────────────────────────────────────────
cosecha_table = []
for cos in sorted(cohort_all, key=sort_key):
    p = parse_cos(cos)
    if not p or p[1] < 2024: continue
    d = cohort_all[cos]
    if d['t'] == 0: continue
    cosecha_table.append({'c':cos,'t':d['t'],'ca':d['ca'],'p':d['p'],'nOrig':N_ORIG.get(cos,0)})

# ── 6. Save data.json ─────────────────────────────────────────────────────────
output = {
    'generated':    datetime.now().strftime('%Y-%m-%dT%H:%M:%SZ'),
    'cosechaTable': cosecha_table,
    'vd':           VD,
    'nOrig':        N_ORIG,
    'origRec':      orig_rec,
    'platTotals':   dict(plat_totals)
}
with open('data.json','w',encoding='utf-8') as f:
    json.dump(output, f, ensure_ascii=False, separators=(',',':'))

n_rec = sum(d['ca'] for d in cohort_all.values())
print(f"✓ data.json generado — {len(cosecha_table)} cohortes 2024+ · {n_rec} recuperaciones · {len(N_ORIG)} meses de originaciones")
