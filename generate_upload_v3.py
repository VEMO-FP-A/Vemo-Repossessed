"""
generate_upload_v3.py  —  VEMO Recuperaciones

USO MAS SIMPLE — dos formas:
  1. Doble clic en el archivo (busca el Excel automáticamente, pide token)
  2. Desde terminal: python generate_upload_v3.py

Flags opcionales:
  --excel   "ruta.xlsx"               (si no se pone, busca automáticamente)
  --token   "ghp_TUTOKEN"             (si no se pone, lo pide por pantalla)
  --repo    "VEMO-FP-A/Vemo-Repossessed"
  --branch  "main"
  --out     "data.json"
  --dry-run  solo genera data.json, NO sube a GitHub
"""

import argparse, json, base64, datetime, sys, os, glob
from collections import defaultdict

# ── Buscar Excel automáticamente ─────────────────────────────────────────────
def find_excel():
    script_dir = os.path.dirname(os.path.abspath(__file__))
    for pattern in ["VEMO_Recuperaciones_Template*.xlsx", "*.xlsx"]:
        matches = glob.glob(os.path.join(script_dir, pattern))
        if matches:
            return matches[0]
    return None

# ── Argumentos ────────────────────────────────────────────────────────────────
def parse_args():
    p = argparse.ArgumentParser()
    p.add_argument("--excel",   default=None)
    p.add_argument("--token",   default=None)
    p.add_argument("--repo",    default="VEMO-FP-A/Vemo-Repossessed")
    p.add_argument("--branch",  default="main")
    p.add_argument("--out",     default="data.json")
    p.add_argument("--dry-run", action="store_true")
    args = p.parse_args()

    if not args.excel:
        args.excel = find_excel()
        if args.excel:
            print(f"Excel encontrado: {os.path.basename(args.excel)}")
        else:
            sys.exit("ERROR: No se encontró ningún .xlsx. Usa --excel 'ruta.xlsx'")

    if not args.dry_run and not args.token:
        args.token = input("GitHub token (ghp_...): ").strip()
        if not args.token:
            print("Sin token — modo dry-run activado automáticamente.")
            args.dry_run = True

    return args

# ── Leer Excel ────────────────────────────────────────────────────────────────
def read_excel(path):
    try:
        import openpyxl
    except ImportError:
        sys.exit("Falta openpyxl. Corre: pip install openpyxl requests")

    wb = openpyxl.load_workbook(path, data_only=True)

    sent_count   = defaultdict(int)
    reposs_count = defaultdict(int)
    cancel_count = defaultdict(int)
    pend_count   = defaultdict(int)
    plat_rec     = defaultdict(int)
    vd_rec_by_m  = defaultdict(lambda: defaultdict(int))

    for row in wb["Export"].iter_rows(min_row=3, values_only=True):
        if not row[0]: continue
        cosecha = str(row[4]).strip() if row[4] else None
        if not cosecha: continue
        estatus  = str(row[7]).strip() if row[7] else ""
        plat     = str(row[11]).strip() if row[11] else ""
        meses_ef = row[12]
        meses_tr = row[6]

        sent_count[cosecha] += 1
        if estatus == "Efectivo":
            reposs_count[cosecha] += 1
            plat_rec[plat] += 1
            try:
                m = int(meses_ef) if meses_ef is not None else int(meses_tr or 0)
            except:
                m = 0
            vd_rec_by_m[cosecha][min(m, 36)] += 1
        elif estatus == "Cancelado":
            cancel_count[cosecha] += 1
        elif estatus == "Pendiente":
            pend_count[cosecha] += 1

    # Vintage curves
    vd = {}
    for cosecha in sorted(set(list(sent_count) + list(vd_rec_by_m))):
        n = sent_count.get(cosecha, 0)
        if n == 0: continue
        cum = [0.0] * 37
        running = 0
        for m in range(37):
            running += vd_rec_by_m[cosecha].get(m, 0)
            cum[m] = round(running / n * 100, 1)
        vd[cosecha] = {"n": n, "c": cum}

    # N_ORIG
    n_orig = {}
    for row in wb["Originaciones"].iter_rows(min_row=3, values_only=True):
        if row[0] and isinstance(row[1], (int, float)):
            n_orig[str(row[0]).strip()] = int(row[1])

    # N_ACTIVA
    n_activa = {}
    if "Flota Activa" in wb.sheetnames:
        for row in wb["Flota Activa"].iter_rows(min_row=3, values_only=True):
            if row[0] and isinstance(row[1], (int, float)):
                n_activa[str(row[0]).strip()] = int(row[1])

    # Cosecha table
    cosecha_table = [
        {
            "c": c,
            "t": sent_count[c],
            "ca": reposs_count.get(c, 0),
            "p": pend_count.get(c, 0),
            "nOrig": n_orig.get(c, sent_count[c]),
            "nActiva": n_activa.get(c, n_orig.get(c, sent_count[c]))
        }
        for c in sorted(sent_count) if sent_count[c] > 0
    ]

    # Orig vs Rec
    MONTHS = ["Jan","Feb","Mar","Apr","May","Jun",
              "Jul","Aug","Sep","Oct","Nov","Dec"]
    orig_rec = {"labels": [], "new_v": [], "rec": []}
    if "Orig_vs_Rec" in wb.sheetnames:
        for row in wb["Orig_vs_Rec"].iter_rows(min_row=3, values_only=True):
            if not row[0]: continue
            mm, yy = str(row[0]).strip().split("-")
            orig_rec["labels"].append(f"{MONTHS[int(mm)-1]} {yy[2:]}")
            orig_rec["new_v"].append(int(row[1]) if row[1] else 0)
            orig_rec["rec"].append(int(row[2]) if row[2] else 0)

    return {
        "vd": vd,
        "nOrig": n_orig,
        "nActiva": n_activa,
        "cosechaTable": cosecha_table,
        "origRec": orig_rec,
        "platTotals": dict(plat_rec),
        "generated": datetime.datetime.utcnow().isoformat() + "Z"
    }

# ── Subir a GitHub ────────────────────────────────────────────────────────────
def upload_to_github(content_str, token, repo, branch, path="data.json"):
    try:
        import requests
    except ImportError:
        sys.exit("Falta requests. Corre: pip install openpyxl requests")

    headers = {
        "Authorization": f"token {token}",
        "Accept": "application/vnd.github.v3+json"
    }
    api = f"https://api.github.com/repos/{repo}/contents/{path}"

    sha = None
    r = requests.get(api, headers=headers, params={"ref": branch})
    if r.status_code == 200:
        sha = r.json().get("sha")
        print("  Archivo existente encontrado, actualizando...")
    elif r.status_code == 404:
        print("  Archivo nuevo, creando...")
    else:
        sys.exit(f"Error al verificar: {r.status_code} {r.text[:200]}")

    payload = {
        "message": f"Update data.json — {datetime.datetime.utcnow().strftime('%Y-%m-%d %H:%M')} UTC",
        "content": base64.b64encode(content_str.encode()).decode(),
        "branch": branch
    }
    if sha:
        payload["sha"] = sha

    r = requests.put(api, headers=headers, json=payload)
    if r.status_code in (200, 201):
        url = r.json().get("content", {}).get("html_url", "")
        print(f"  Subido exitosamente!")
        print(f"  URL: {url}")
    else:
        sys.exit(f"Error al subir: {r.status_code} {r.text[:300]}")

# ── Main ──────────────────────────────────────────────────────────────────────
if __name__ == "__main__":
    args = parse_args()

    print(f"\nLeyendo {os.path.basename(args.excel)}...")
    data = read_excel(args.excel)

    print(f"  Cosechas:    {len(data['vd'])}")
    print(f"  Tabla:       {len(data['cosechaTable'])} filas")
    print(f"  Orig/Rec:    {len(data['origRec']['labels'])} meses")
    print(f"  Plataformas: {data['platTotals']}")

    # Guardar data.json en la misma carpeta que el script
    out_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), args.out)
    content_str = json.dumps(data, ensure_ascii=False, indent=2)
    with open(out_path, "w", encoding="utf-8") as f:
        f.write(content_str)
    print(f"\nGuardado: {args.out} ({len(content_str):,} bytes)")

    if not args.dry_run:
        print(f"\nSubiendo a {args.repo} ({args.branch})...")
        upload_to_github(content_str, args.token, args.repo, args.branch)
    else:
        print("Modo dry-run — no se subió a GitHub.")

    input("\nPresiona Enter para cerrar...")
